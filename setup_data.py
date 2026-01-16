#!/usr/bin/env python3
"""
StudyHub Data Setup Script
==========================

This script helps you:
1. Create a properly formatted Google Sheet from the Excel template
2. Convert existing data to the required format
3. Validate your data structure
4. Generate sample data for testing

Usage:
    python setup_data.py --help
    python setup_data.py create-sample
    python setup_data.py validate path/to/data.xlsx
    python setup_data.py export-json path/to/data.xlsx
"""

import json
import os
import sys
from datetime import datetime

# Check for required packages
try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
except ImportError:
    print("Installing required packages...")
    os.system(f"{sys.executable} -m pip install pandas openpyxl --break-system-packages")
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ============================================================================
# DATA STRUCTURE DEFINITIONS
# ============================================================================

SHEET_SCHEMAS = {
    'Subjects': {
        'columns': ['subject_id', 'subject_key', 'name', 'icon', 'color_hex', 'light_bg', 'gradient_from', 'gradient_to', 'dark_glow'],
        'required': ['subject_id', 'subject_key', 'name'],
        'description': 'Define subjects with their visual styling'
    },
    'Topics': {
        'columns': ['topic_id', 'subject_key', 'topic_name', 'duration_minutes', 'order_index'],
        'required': ['topic_id', 'subject_key', 'topic_name'],
        'description': 'List all topics per subject'
    },
    'Topic_Sections': {
        'columns': ['section_id', 'topic_id', 'section_title', 'section_icon', 'order_index', 'section_type'],
        'required': ['section_id', 'topic_id', 'section_title'],
        'description': 'Define sections/chapters within each topic'
    },
    'Learning_Objectives': {
        'columns': ['objective_id', 'topic_id', 'objective_text', 'order_index'],
        'required': ['objective_id', 'topic_id', 'objective_text'],
        'description': 'Learning objectives for each topic'
    },
    'Key_Terms': {
        'columns': ['term_id', 'topic_id', 'term', 'definition'],
        'required': ['term_id', 'topic_id', 'term', 'definition'],
        'description': 'Vocabulary terms and definitions'
    },
    'Study_Content': {
        'columns': ['content_id', 'section_id', 'content_type', 'content_title', 'content_text', 'order_index'],
        'required': ['content_id', 'section_id', 'content_type', 'content_text'],
        'description': 'Main educational content blocks'
    },
    'Formulas': {
        'columns': ['formula_id', 'topic_id', 'formula_text', 'formula_label', 
                   'variable_1_symbol', 'variable_1_name', 'variable_1_unit',
                   'variable_2_symbol', 'variable_2_name', 'variable_2_unit',
                   'variable_3_symbol', 'variable_3_name', 'variable_3_unit'],
        'required': ['formula_id', 'topic_id', 'formula_text'],
        'description': 'Mathematical/scientific formulas'
    },
    'Quiz_Questions': {
        'columns': ['question_id', 'topic_id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 
                   'correct_answer', 'explanation', 'xp_reward'],
        'required': ['question_id', 'topic_id', 'question_text', 'option_a', 'option_b', 'correct_answer'],
        'description': 'Multiple choice quiz questions'
    },
    'Achievements': {
        'columns': ['achievement_id', 'icon', 'name', 'description', 'unlock_condition'],
        'required': ['achievement_id', 'name', 'description'],
        'description': 'Gamification badges and achievements'
    }
}

CONTENT_TYPES = ['introduction', 'formula', 'concept_helper', 'warning', 'real_world', 'text']
SECTION_TYPES = ['objectives', 'intro', 'content', 'applications', 'quiz']
VALID_ICONS = ['Zap', 'Calculator', 'FlaskConical', 'Leaf', 'Trophy', 'Star', 'Award', 'Flame',
               'HelpCircle', 'CheckCircle2', 'Target', 'BookOpen', 'FileText', 'Clock', 'Globe',
               'Lightbulb', 'AlertTriangle']


# ============================================================================
# SAMPLE DATA
# ============================================================================

SAMPLE_DATA = {
    'Subjects': [
        {'subject_id': 'phys-001', 'subject_key': 'physics', 'name': 'Physics', 'icon': 'Zap', 
         'color_hex': '#3B82F6', 'light_bg': 'bg-blue-50', 'gradient_from': 'blue-500', 
         'gradient_to': 'blue-600', 'dark_glow': 'shadow-blue-500/20'},
        {'subject_id': 'math-001', 'subject_key': 'math', 'name': 'Mathematics', 'icon': 'Calculator',
         'color_hex': '#10B981', 'light_bg': 'bg-emerald-50', 'gradient_from': 'emerald-500',
         'gradient_to': 'emerald-600', 'dark_glow': 'shadow-emerald-500/20'},
        {'subject_id': 'chem-001', 'subject_key': 'chemistry', 'name': 'Chemistry', 'icon': 'FlaskConical',
         'color_hex': '#F59E0B', 'light_bg': 'bg-amber-50', 'gradient_from': 'amber-500',
         'gradient_to': 'amber-600', 'dark_glow': 'shadow-amber-500/20'},
        {'subject_id': 'bio-001', 'subject_key': 'biology', 'name': 'Biology', 'icon': 'Leaf',
         'color_hex': '#8B5CF6', 'light_bg': 'bg-violet-50', 'gradient_from': 'violet-500',
         'gradient_to': 'violet-600', 'dark_glow': 'shadow-violet-500/20'},
    ],
    'Topics': [
        {'topic_id': 'phys-t001', 'subject_key': 'physics', 'topic_name': "Newton's Laws of Motion", 'duration_minutes': 25, 'order_index': 1},
        {'topic_id': 'phys-t002', 'subject_key': 'physics', 'topic_name': 'Work and Energy', 'duration_minutes': 30, 'order_index': 2},
        {'topic_id': 'phys-t003', 'subject_key': 'physics', 'topic_name': 'Light and Optics', 'duration_minutes': 20, 'order_index': 3},
        {'topic_id': 'math-t001', 'subject_key': 'math', 'topic_name': 'Exponents', 'duration_minutes': 20, 'order_index': 1},
        {'topic_id': 'math-t002', 'subject_key': 'math', 'topic_name': 'Probability', 'duration_minutes': 25, 'order_index': 2},
        {'topic_id': 'chem-t001', 'subject_key': 'chemistry', 'topic_name': 'Atomic Structure', 'duration_minutes': 25, 'order_index': 1},
        {'topic_id': 'bio-t001', 'subject_key': 'biology', 'topic_name': 'Cell Structure', 'duration_minutes': 25, 'order_index': 1},
    ],
    'Topic_Sections': [
        {'section_id': 'phys-t001-s001', 'topic_id': 'phys-t001', 'section_title': 'Learning Objectives', 'section_icon': 'Target', 'order_index': 1, 'section_type': 'objectives'},
        {'section_id': 'phys-t001-s002', 'topic_id': 'phys-t001', 'section_title': 'Introduction', 'section_icon': 'BookOpen', 'order_index': 2, 'section_type': 'intro'},
        {'section_id': 'phys-t001-s003', 'topic_id': 'phys-t001', 'section_title': "Newton's First Law", 'section_icon': 'Zap', 'order_index': 3, 'section_type': 'content'},
        {'section_id': 'phys-t001-s004', 'topic_id': 'phys-t001', 'section_title': "Newton's Second Law", 'section_icon': 'Zap', 'order_index': 4, 'section_type': 'content'},
        {'section_id': 'phys-t001-s005', 'topic_id': 'phys-t001', 'section_title': "Newton's Third Law", 'section_icon': 'Zap', 'order_index': 5, 'section_type': 'content'},
        {'section_id': 'phys-t001-s006', 'topic_id': 'phys-t001', 'section_title': 'Real-World Applications', 'section_icon': 'Globe', 'order_index': 6, 'section_type': 'applications'},
        {'section_id': 'phys-t001-s007', 'topic_id': 'phys-t001', 'section_title': 'Topic Quiz', 'section_icon': 'HelpCircle', 'order_index': 7, 'section_type': 'quiz'},
    ],
    'Learning_Objectives': [
        {'objective_id': 'obj-phys-t001-01', 'topic_id': 'phys-t001', 'objective_text': 'Explain the concept of inertia and how it relates to mass', 'order_index': 1},
        {'objective_id': 'obj-phys-t001-02', 'topic_id': 'phys-t001', 'objective_text': 'Apply the formula F = ma to solve real-world problems', 'order_index': 2},
        {'objective_id': 'obj-phys-t001-03', 'topic_id': 'phys-t001', 'objective_text': 'Identify action-reaction force pairs in various scenarios', 'order_index': 3},
        {'objective_id': 'obj-phys-t001-04', 'topic_id': 'phys-t001', 'objective_text': "Analyze motion using all three of Newton's Laws", 'order_index': 4},
    ],
    'Key_Terms': [
        {'term_id': 'term-phys-t001-01', 'topic_id': 'phys-t001', 'term': 'Force', 'definition': 'A push or pull on an object'},
        {'term_id': 'term-phys-t001-02', 'topic_id': 'phys-t001', 'term': 'Mass', 'definition': 'Amount of matter in an object (kg)'},
        {'term_id': 'term-phys-t001-03', 'topic_id': 'phys-t001', 'term': 'Acceleration', 'definition': 'Rate of change of velocity (m/s²)'},
        {'term_id': 'term-phys-t001-04', 'topic_id': 'phys-t001', 'term': 'Inertia', 'definition': 'Resistance to change in motion'},
        {'term_id': 'term-phys-t001-05', 'topic_id': 'phys-t001', 'term': 'Newton (N)', 'definition': 'SI unit of force (kg·m/s²)'},
    ],
    'Study_Content': [
        {'content_id': 'cont-001', 'section_id': 'phys-t001-s004', 'content_type': 'introduction', 'content_title': 'Introduction',
         'content_text': "Newton's Second Law describes what happens when an unbalanced force acts on an object. It quantifies the relationship between force, mass, and acceleration.", 'order_index': 1},
        {'content_id': 'cont-002', 'section_id': 'phys-t001-s004', 'content_type': 'formula', 'content_title': 'The Formula',
         'content_text': 'F = m × a', 'order_index': 2},
        {'content_id': 'cont-003', 'section_id': 'phys-t001-s004', 'content_type': 'concept_helper', 'content_title': 'Concept Helper',
         'content_text': 'Think of it like pushing a shopping cart. An empty cart (less mass) accelerates quickly with a small push. A full cart (more mass) needs more force for the same acceleration!', 'order_index': 3},
        {'content_id': 'cont-004', 'section_id': 'phys-t001-s004', 'content_type': 'warning', 'content_title': 'Common Misunderstanding',
         'content_text': 'Students often confuse mass and weight. Mass is the amount of matter (measured in kg) and stays constant. Weight is the force of gravity on that mass (measured in N) and changes based on location!', 'order_index': 4},
        {'content_id': 'cont-005', 'section_id': 'phys-t001-s004', 'content_type': 'real_world', 'content_title': 'Real-World Application',
         'content_text': 'Car engineers use F = ma to calculate braking distances. More massive vehicles need stronger brakes!', 'order_index': 5},
    ],
    'Formulas': [
        {'formula_id': 'formula-001', 'topic_id': 'phys-t001', 'formula_text': 'F = m × a', 'formula_label': "Newton's Second Law",
         'variable_1_symbol': 'F', 'variable_1_name': 'Force', 'variable_1_unit': 'N',
         'variable_2_symbol': 'm', 'variable_2_name': 'Mass', 'variable_2_unit': 'kg',
         'variable_3_symbol': 'a', 'variable_3_name': 'Acceleration', 'variable_3_unit': 'm/s²'},
    ],
    'Quiz_Questions': [
        {'question_id': 'quiz-phys-t001-01', 'topic_id': 'phys-t001', 
         'question_text': 'If a 10 kg object accelerates at 2 m/s², what is the force?',
         'option_a': '5 N', 'option_b': '20 N', 'option_c': '12 N', 'option_d': '8 N',
         'correct_answer': 'B', 'explanation': 'Using F = m × a: F = 10 × 2 = 20 N', 'xp_reward': 10},
        {'question_id': 'quiz-phys-t001-02', 'topic_id': 'phys-t001',
         'question_text': "Newton's First Law is also known as the law of:",
         'option_a': 'Acceleration', 'option_b': 'Action-Reaction', 'option_c': 'Inertia', 'option_d': 'Gravity',
         'correct_answer': 'C', 'explanation': "Newton's First Law describes inertia - objects resist changes in motion", 'xp_reward': 10},
        {'question_id': 'quiz-phys-t001-03', 'topic_id': 'phys-t001',
         'question_text': "Which statement best describes Newton's Third Law?",
         'option_a': 'F = ma', 'option_b': 'Objects at rest stay at rest', 
         'option_c': 'Every action has an equal and opposite reaction', 'option_d': 'Heavier objects fall faster',
         'correct_answer': 'C', 'explanation': "Newton's Third Law states that forces come in pairs", 'xp_reward': 10},
    ],
    'Achievements': [
        {'achievement_id': 'first-login', 'icon': 'Zap', 'name': 'First Login', 'description': 'Welcome to StudyHub!', 'unlock_condition': 'Login for the first time'},
        {'achievement_id': 'first-quiz', 'icon': 'HelpCircle', 'name': 'First Quiz', 'description': 'Complete your first quiz', 'unlock_condition': 'Complete any quiz'},
        {'achievement_id': 'streak-5', 'icon': 'Flame', 'name': '5-Day Streak', 'description': 'Study 5 days in a row', 'unlock_condition': 'streak >= 5'},
        {'achievement_id': 'streak-10', 'icon': 'Flame', 'name': '10-Day Streak', 'description': 'Study 10 days in a row', 'unlock_condition': 'streak >= 10'},
        {'achievement_id': 'topic-complete', 'icon': 'CheckCircle2', 'name': 'Topic Master', 'description': 'Complete any topic', 'unlock_condition': 'Any topic progress = 100'},
        {'achievement_id': 'subject-50', 'icon': 'Trophy', 'name': 'Halfway There', 'description': '50% in any subject', 'unlock_condition': 'Any subject progress >= 50'},
        {'achievement_id': 'perfect-quiz', 'icon': 'Star', 'name': 'Perfect Score', 'description': 'Score 100% on a quiz', 'unlock_condition': 'Any quiz score = 100'},
        {'achievement_id': 'all-subjects', 'icon': 'Award', 'name': 'Well Rounded', 'description': 'Study all 4 subjects', 'unlock_condition': 'All subjects accessed'},
    ]
}


# ============================================================================
# FUNCTIONS
# ============================================================================

def create_sample_excel(output_path='StudyHub_Sample_Data.xlsx'):
    """Create a sample Excel file with all the required sheets and data."""
    print(f"Creating sample Excel file: {output_path}")
    
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for i, (sheet_name, data) in enumerate(SAMPLE_DATA.items()):
        if i == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        
        schema = SHEET_SCHEMAS.get(sheet_name, {})
        columns = schema.get('columns', list(data[0].keys()) if data else [])
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = max(len(str(col_name)), max(len(str(row.get(col_name, ''))) for row in data) if data else 0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    wb.save(output_path)
    print(f"✅ Sample Excel file created: {output_path}")
    print(f"   Sheets created: {', '.join(SAMPLE_DATA.keys())}")
    return output_path


def validate_excel(file_path):
    """Validate an Excel file against the required schema."""
    print(f"Validating: {file_path}")
    errors = []
    warnings = []
    
    try:
        wb = load_workbook(file_path)
    except Exception as e:
        print(f"❌ Error opening file: {e}")
        return False
    
    # Check for required sheets
    for sheet_name, schema in SHEET_SCHEMAS.items():
        if sheet_name not in wb.sheetnames:
            errors.append(f"Missing required sheet: {sheet_name}")
            continue
        
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1] if cell.value]
        headers_lower = [h.lower().replace(' ', '_') if h else '' for h in headers]
        
        # Check required columns
        for required_col in schema.get('required', []):
            if required_col not in headers_lower:
                errors.append(f"{sheet_name}: Missing required column '{required_col}'")
        
        # Check for data
        if ws.max_row < 2:
            warnings.append(f"{sheet_name}: No data rows found")
        
        # Validate content types for Study_Content
        if sheet_name == 'Study_Content' and 'content_type' in headers_lower:
            type_col = headers_lower.index('content_type') + 1
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=type_col).value
                if cell_value and cell_value not in CONTENT_TYPES:
                    warnings.append(f"{sheet_name} row {row}: Invalid content_type '{cell_value}'. Valid: {CONTENT_TYPES}")
        
        # Validate icons
        if sheet_name in ['Subjects', 'Topic_Sections', 'Achievements']:
            icon_col_name = 'icon' if sheet_name != 'Topic_Sections' else 'section_icon'
            if icon_col_name in headers_lower:
                icon_col = headers_lower.index(icon_col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=icon_col).value
                    if cell_value and cell_value not in VALID_ICONS:
                        warnings.append(f"{sheet_name} row {row}: Unknown icon '{cell_value}'")
    
    # Print results
    print("\n" + "="*50)
    if errors:
        print("❌ ERRORS:")
        for error in errors:
            print(f"   • {error}")
    
    if warnings:
        print("⚠️  WARNINGS:")
        for warning in warnings:
            print(f"   • {warning}")
    
    if not errors and not warnings:
        print("✅ Validation passed! No issues found.")
    elif not errors:
        print("✅ Validation passed with warnings.")
    else:
        print("❌ Validation failed. Please fix errors before using.")
    
    print("="*50)
    return len(errors) == 0


def export_to_json(file_path, output_path=None):
    """Export Excel data to JSON format for use without Google Sheets."""
    print(f"Exporting to JSON: {file_path}")
    
    if output_path is None:
        output_path = file_path.replace('.xlsx', '.json')
    
    wb = load_workbook(file_path)
    data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers
        headers = [cell.value.lower().replace(' ', '_') if cell.value else f'col_{i}' 
                   for i, cell in enumerate(ws[1])]
        
        # Get data
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value if value is not None else ''
                rows.append(row_dict)
        
        data[sheet_name] = rows
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"✅ JSON exported: {output_path}")
    return output_path


def print_schema():
    """Print the data schema for reference."""
    print("\n" + "="*60)
    print("STUDYHUB DATA SCHEMA")
    print("="*60)
    
    for sheet_name, schema in SHEET_SCHEMAS.items():
        print(f"\n📄 {sheet_name}")
        print(f"   Description: {schema['description']}")
        print(f"   Columns: {', '.join(schema['columns'])}")
        print(f"   Required: {', '.join(schema['required'])}")
    
    print("\n" + "-"*60)
    print("VALID CONTENT TYPES:", ', '.join(CONTENT_TYPES))
    print("VALID SECTION TYPES:", ', '.join(SECTION_TYPES))
    print("VALID ICONS:", ', '.join(VALID_ICONS))
    print("="*60)


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='StudyHub Data Setup Script')
    parser.add_argument('command', choices=['create-sample', 'validate', 'export-json', 'schema'],
                       help='Command to run')
    parser.add_argument('file', nargs='?', help='Input file path (for validate/export-json)')
    parser.add_argument('-o', '--output', help='Output file path')
    
    args = parser.parse_args()
    
    if args.command == 'create-sample':
        output = args.output or 'StudyHub_Sample_Data.xlsx'
        create_sample_excel(output)
    
    elif args.command == 'validate':
        if not args.file:
            print("Error: Please provide a file path to validate")
            sys.exit(1)
        success = validate_excel(args.file)
        sys.exit(0 if success else 1)
    
    elif args.command == 'export-json':
        if not args.file:
            print("Error: Please provide a file path to export")
            sys.exit(1)
        export_to_json(args.file, args.output)
    
    elif args.command == 'schema':
        print_schema()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nQuick start:")
        print("  python setup_data.py create-sample   # Create sample Excel file")
        print("  python setup_data.py schema          # Show data schema")
    else:
        main()
